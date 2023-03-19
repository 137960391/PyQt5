import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import xlrd
import UITABLE
from TYPE import FileType, ReplaceData
from datetime import datetime
import os
import configparser

dpath = os.path.join(os.path.expanduser("~"), "Desktop")
dpath = dpath.replace("\\", "/")
dpath = dpath + '/每日销售情况最新.xlsx'


class DataInfo:
    def __init__(self, path1, path2, datatime, sale_money_text):
        self.ut = UITABLE.UiTable(datatime, sale_money_text)
        self.datatime = datatime
        self.title_row = self.ut.title_row
        self.title_col = ''
        self.title_col = self.ut.title_col
        self.head = self.ut.head
        self.data1 = []  # 合并得到汇总数据
        self.data2 = []  # 数据源中的类别判断语句
        self.data3 = [[], []]  # 加入表格中数据
        self.title_col_list = self.get_title_col_list()
        self.workbook2 = openpyxl.load_workbook(path2, data_only=True, read_only=False) # data_only只读数据，不读公式
        self.workbook1 = xlrd.open_workbook(path1)
        self.deal_excel_data(path1, path2)
        self.mydict = []
        self.rreplace = []
        for x in self.title_col_list:
            self.mydict = self.mydict + [FileType(x)]
        self.total_money = 0
        self.total_count = 0
        self.get_title_col_list()
        self.total()

    def get_title_col_list(self):
        start_index = self.title_col.index('产品')
        end_index = self.title_col.index('汇总')
        return self.title_col[start_index + 1:end_index]

    # 将data1 转为其他
    def find_second_column_value(self):
        for num, x in enumerate(self.data1):
            temp = self.data1[num]
            for row in self.data2:
                if row[0] == x[0] and row[1] in self.title_col_list:
                    self.data1[num] = self.data1[num] + [row[1]]
                    break
            if temp == self.data1[num]:
                self.data1[num] = self.data1[num] + ['其他文具']
            self.total_money = round(self.total_money + x[2], 2)
            self.total_count = round(self.total_count + x[1], 2)
        return None

    def get_column_data_xlsx(self, sheet, start_row, start_col1, start_col2):
        # 0-indexed
        start_col1 = start_col1 - 1   # 0-indexed
        start_col2 = start_col2 - 1   # 0-indexed
        _data = []
        self.rreplace = []
        config = configparser.ConfigParser()
        config.read('../../config.ini')
        for key in config['replace_data']:
            self.rreplace = self.rreplace + [ReplaceData(self.parse_config(config['replace_data'][key]))]
        for row in sheet.iter_rows(min_row=start_row+1, values_only=True):
            cell1 = int(row[start_col1])
            cell2 = row[start_col2]
            if self.rreplace:
                for x in self.rreplace:
                    if cell2 in x.replace_name_before:
                        cell2 = x.replace_name[0]
            _data.append([cell1, cell2])
        return _data

    @staticmethod
    def parse_config(config_string):
        # key_value_pair = config_string.split('=')
        # key = key_value_pair[0].strip()
        # value = key_value_pair[1].strip()
        value_list = [item.strip() for item in config_string.split(',')]
        first_item_split = value_list[0].split(' ')
        result = [first_item_split] + [value_list[1:]]

        return result

    @staticmethod
    def get_column_data_xls(sheet, start_row, start_col1, start_col2, start_col3):
        # 0-indexed
        start_col1 = start_col1 - 1   # 0-indexed
        start_col2 = start_col2 - 1   # 0-indexed
        start_col3 = start_col3 - 1  # 0-indexed
        # data = sheet.cell_value(start_row, start_col1)
        _data = []
        row = start_row
        while row < sheet.nrows:
            cell1 = int(sheet.cell_value(row, start_col1))
            cell2 = int(sheet.cell_value(row, start_col2))
            cell3 = round(float(sheet.cell_value(row, start_col3).replace(',', '')), 2)
            _data = _data + [[cell1, cell2, cell3]]
            row += 1

        return _data

    @staticmethod
    def search_in_excel_xls(sheet, search_value):
        # 遍历单元格，查找匹配值
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                if sheet.cell(row, col).value == search_value:
                    # 找到匹配值，返回行列
                    return row + 1, col + 1
        # 未找到匹配值，返回 None

        return None

    @staticmethod
    def search_in_excel_xlsx(sheet, search_value):
        # 遍历单元格，查找匹配值
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_idx, col in enumerate(row, start=1):
                if col == search_value:
                    # 找到匹配值，返回行列
                    return row_idx, col_idx
        # 未找到匹配值，返回 None

        return None

    @staticmethod
    def get_excel_type(filename):

        if filename.endswith('.xlsx'):
            return '.xlsx'
        elif filename.endswith('.xls'):
            return '.xls'
        else:
            return -1

    def open_excel_xls(self):

        # 获取第一个表格
        worksheet = self.workbook1.sheet_by_index(0)
        _col = self.search_in_excel_xls(worksheet, '商品ID')[0]
        _id = self.search_in_excel_xls(worksheet, '商品ID')[1]
        _count = self.search_in_excel_xls(worksheet, '支付件数')[1]
        _money = self.search_in_excel_xls(worksheet, '支付金额')[1]
        return self.get_column_data_xls(worksheet, _col, _id, _count, _money)

    def open_excel_xlsx(self):

        worksheet = self.workbook2['数据源']
        _col = self.search_in_excel_xlsx(worksheet, '商品ID')[0]
        _id = self.search_in_excel_xlsx(worksheet, '商品ID')[1]
        _count = self.search_in_excel_xlsx(worksheet, '类别')[1]
        return self.get_column_data_xlsx(worksheet, _col, _id, _count)

    def deal_excel_data(self, book1, book2):

        if self.get_excel_type(book1) == '.xls':
            self.data1 = self.open_excel_xls()
        else:
            self.data1 = self.open_excel_xlsx()

        if self.get_excel_type(book2) == 'xls':
            self.data2 = self.open_excel_xls()
        else:
            self.data2 = self.open_excel_xlsx()

    def total(self):

        # 得到data3数据填充
        self.find_second_column_value()
        for x in self.data1:
            for y in self.mydict:
                if x[3] == y.type:
                    y.count = y.count + x[1]
                    y.money = y.money + round(x[2], 2)
        for x in self.mydict:
            self.data3[0] = self.data3[0] + [x.count]
            self.data3[1] = self.data3[1] + [x.money]

        # 设置字体和字号
        font = Font(name='微软雅黑', size=10)
        # 设置边框
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        sheet = self.workbook2['每日销售情况']
        p_count = sheet.max_row
        # 设置第 1 行第 1 列到第 3 列的单元格合并
        cell = sheet.merge_cells(start_row=p_count + 4, start_column=1, end_row=p_count + 4, end_column=5)
        # 在首行合并的单元格中写入数据
        for x in range(5):
            cell = sheet.cell(row=p_count + 4, column=x+1)
            cell.font = font
            cell.border = border
        cell = sheet.cell(row=p_count + 4, column=1, value=self.head)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 表格加边框
        for num1, x in enumerate(self.title_col):
            for num2, y in enumerate(self.title_row):
                cell = sheet.cell(row=p_count + 5 + num1, column=num2 + 1)
                cell.font = font
                cell.border = border
        # # 底部合并
        # cell = sheet.merge_cells(start_row=p_count + len(self.title_col), start_column=p_count + len(self.title_col),
        #                          end_row=p_count + 4, end_column=5)
        # 类别
        for num, x in enumerate(self.title_row):
            sheet.cell(row=p_count + 5, column=num + 1, value=x)
        # 产品具体类

        count = 0
        for num, x in enumerate(self.title_col):
            sheet.cell(row=p_count + 5 + num, column=1, value=x)
            if x in self.title_col_list:
                sheet.cell(row=p_count + 5 + num, column=2, value=self.data3[0][count])
                sheet.cell(row=p_count + 5 + num, column=3,
                           value='{:.2f}%'.format(self.data3[0][count]/self.total_count*100))\
                        .alignment = Alignment(horizontal='right', vertical='center')
                sheet.cell(row=p_count + 5 + num, column=4, value=self.data3[1][count])
                sheet.cell(row=p_count + 5 + num, column=5,
                           value='{:.2f}%'.format(self.data3[1][count]/self.total_money*100))\
                        .alignment = Alignment(horizontal='right', vertical='center')
                count = count + 1
        sheet.cell(row=p_count + 6 + count, column=2, value=self.total_count)
        sheet.cell(row=p_count + 6 + count, column=4, value=self.total_money)
        sheet.merge_cells(start_row=p_count + 7 + count, start_column=2, end_row=p_count + 7 + count, end_column=5)
        # 写入累加值

        cell = sheet.cell(row=p_count + 7 + count, column=2)
        while 1:
            if self.get_before_data(sheet, p_count) == -1:
                cell.value = self.total_money
                break
            if int(self.get_before_data(sheet, p_count)) != int(self.get_month(self.datatime)):
                cell.value = self.total_money + 0

                break
            else:
                cell.value = sheet.cell(row=p_count, column=2).value + self.total_money
                break
        cell.alignment = Alignment(horizontal='center', vertical='center')
        '''
        [654885549470, 22, 1432.98, '国产相册'], [656447053097, 16, 1164.04, '墨水']
        '''

        self.workbook2.save(dpath)

        exit(0)
        pass

    def get_before_data(self, sheet, p):
        temp = p
        if p < 10:
            return -1
        text = ''
        while p != temp-80:
            cell = sheet.cell(row=p, column=1)
            if '销售明细' in cell.value:
                text = cell.value
                break
            p = p - 1
        return self.extract_month_number(text)

    @staticmethod
    def extract_month_number(s):
        index = s.find('月')  # 找到 '月' 字符的索引
        if index > 0:
            i = index - 1
            while i >= 0 and s[i].isdigit():
                i -= 1
            return s[i + 1:index]
        else:
            return None

    @staticmethod
    def get_month(date_str):
        # 解析日期字符串为日期对象
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        # 获取日期对象的月份并返回
        return date_obj.month


if __name__ == '__main__':

    data = DataInfo('C:/Users/DH LOVE CHQ/Desktop/【生意参谋平台】商品_全部_2023-03-13_2023-03-13 (1)(1).xls',
                    'C:/Users/DH LOVE CHQ/Desktop/每日销售情况最新(1).xlsx', '2023-12-01', '2000')
    # data.total()

